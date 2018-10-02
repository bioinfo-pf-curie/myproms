#! /bioinfo/local/build/python/python-2.7.9/bin/python

################################################################################
# convert_xlsx2txt.py      1.0.1                                               #
# Authors: P. Poullet (Institut Curie)                                         #
# Contact: myproms@curie.fr                                                    #
# Exports xlsx file to txt: tab-separated                                      #
################################################################################
#----------------------------------CeCILL License-------------------------------
# This file is part of myProMS
#
# Copyright Institut Curie 2018
#
# This software is a computer program whose purpose is to process
# Mass Spectrometry-based proteomic data.
#
# This software is governed by the CeCILL license under French law and
# abiding by the rules of distribution of free software. You can use,
# modify and/or redistribute the software under the terms of the CeCILL
# license as circulated by CEA, CNRS and INRIA at the following URL
# "http://www.cecill.info".
#
# As a counterpart to the access to the source code and rights to copy,
# modify and redistribute granted by the license, users are provided only
# with a limited warranty and the software's author, the holder of the
# economic rights, and the successive licensors have only limited
# liability.
#
# In this respect, the user's attention is drawn to the risks associated
# with loading, using, modifying and/or developing or reproducing the
# software by the user in light of its specific status of free software,
# that may mean that it is complicated to manipulate, and that also
# therefore means that it is reserved for developers and experienced
# professionals having in-depth computer knowledge. Users are therefore
# encouraged to load and test the software's suitability as regards their
# requirements in conditions enabling the security of their systems and/or
# data to be ensured and, more generally, to use and operate it in the
# same conditions as regards security.
#
# The fact that you are presently reading this means that you have had
# knowledge of the CeCILL license and that you accept its terms.
#-------------------------------------------------------------------------------

import sys, getopt, re
#import os
from openpyxl import load_workbook


def checkArguments(argv):
	argMsg='Usage: convert_xlsx2tsv.py -f <Excel file> [-s <Sheet ranks (comma-separated)>]'
	inputFile=''
	sheetRanks=[]
	try:
		opts, args = getopt.getopt(argv,"hf:s:",["help","file=","sheets="])
		if len(opts) == 0:
			print argMsg
			sys.exit()
	except getopt.GetoptError:
		print argMsg
		sys.exit()

	for opt, arg in opts:
		if opt == '-h' or opt == '--help':
			print argMsg
			sys.exit()
		elif opt in ("-f", "--file"):
			inputFile = arg
		elif opt in ("-s", "--sheets"):
			sheetRanks=arg.split(',')
	return inputFile,sheetRanks


inFile,sheetList=checkArguments(sys.argv[1:])
#sys.exit()

#inFileInfo=os.path.split(inFile)
inFileRootName=re.sub(r'\.xlsx$','',inFile)
#print inFileRootName
try:
	wb = load_workbook(filename=inFile, read_only=True)
except IOError:
	print 'ERROR: Could not load file "'+inFile+'"!'
	sys.exit()

sh=open(inFileRootName + '_SheetNames.info','w')

sheetRank=0
for wsName in wb.get_sheet_names():
	sheetRank+=1
	if len(sheetList) and str(sheetRank) not in (sheetList):
		continue

	ws = wb[wsName] # ws is now an IterableWorksheet
	if ws.max_row==1: # empty sheet
		continue
	print('Exporting sheet ['+wsName+']...')

	wsNameMod=re.sub(r'\W', '_', wsName)

	outFile=inFileRootName + '-' + wsNameMod + '.tsv'
	sh.write(wsName + "\t" + outFile + "\n")
	out=open(outFile, 'w')

	#rowCount=0;
	for row in ws.rows:
		rowData=[]
		for cell in row:
			try:
				rowData.append(str(cell.value))
			except:
				rowData.append('')
		#rowCount+=1
		#if rowCount==10:
		#	break
		out.write("\t".join(rowData) + "\n")
		#print rowData
	out.close()
	print(' Done.')


sh.close()

####>Revision history<####
# 1.0.1 Improved argument error checks (PP 13/07/16)
# 1.0.0 Exports to tab-separated .tsv files (PP 06/05/16)
